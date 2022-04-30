from openpyxl import load_workbook
from pathlib import Path
from docxtpl import DocxTemplate
import argparse
import shutil
import email, smtplib, ssl
from email import encoders
import yagmail
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

""" scheduler.py służy do tworzenia raportów dla inspekcji.
Program wywołuje się wpisując w terminalu 
    python3 scheduler.py source template
przykładowo
    python3 scheduler.py odbiory.xlsx formatka_A35
wygenerowane raporty można znaleźć w folderze OUTPUT
"""

parser = argparse.ArgumentParser(description='wypluwa raporty dla odbiorow')
parser.add_argument('source', metavar='zrodlo', type=str, help='source   czyli źródłowy plik excel z odbiorami zgłoszonymi do Calypso')
parser.add_argument('template', metavar='formatka', type=str, help='template   czyli plik służący za formatkę dla danego projektu stoczniowego')
args = parser.parse_args()

source = args.source
template = args.template



base_dir = Path(__file__).parent
workbook = load_workbook(filename=base_dir / source)
sheet = workbook.active

word_template_path = base_dir / template
doc = DocxTemplate(word_template_path)
output_dir = base_dir / "OUTPUT"
output_dir.mkdir(exist_ok=True)


for row in sheet.iter_rows(min_row=2,values_only=True):
    Inspection_ID = row[1]
    Day = row[3]
    Starting_time = row[4]
    Description = row[15]
    Inspector = row[10]
    JanLuczewski = "JAN LUCZEWSKI"
    if Inspector == JanLuczewski:
        context = {
            "Inspection_ID" : Inspection_ID,
            "Description" : Description,
            "Starting_time" : Starting_time,
            "Day" : Day,
        }
        doc.render(context)
        doc.save(output_dir / f"{Inspection_ID}.docx")
shutil.make_archive(source,'zip','OUTPUT')

sender = 'jan.luczewski@gmail.com'
receiver = 'jan.luczewski@crist.com.pl'
text = """\

pozdrowionka
"""

html = """\
<html>
    <body>
        <h1>Oto papiery </h1>
        <h3>załącznik podsyłam w formacie zip</h3>
        <h4>zapisz załącznik</h4>
        <h4>najedź myszą na plik w zapisanym folderze, użyj 7-zip aby rozpakować</h4>
        <h2>Pozdrawiam ,wracam w piątek</h2>     
    </body>
</html>
"""

filename = f'{source}.zip'
yagmail.register('jan.luczewski@gmail.com','paniwladca')
yag = yagmail.SMTP('jan.luczewski@gmail.com')
yag.send(to=receiver,
         subject='Odbiory3',
         contents=[text,html],
         attachments=filename)

print('wysłano')
